"""Ingest GPS/wearable training data from Catapult-style XLSX exports.

Each XLSX has one sheet with 17 columns: session, week, player, distance,
work-rate, duration, 4 accel zones, 2 high-intensity (acc/dec), 3 speed zones,
power_avg, sprints_per_min.

Player short-names in the XLSX are matched to Player rows via an alias map.
"""
import re
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from insights.models import Player, TrainingSession


# Maps XLSX short-names (column "Players") to a substring of Player.name in DB.
# `None` means we intentionally skip (U21 / test players we don't have Wyscout data for).
ALIAS = {
    'Tosca':      'Toșca',
    'Lukic':      'Lukić',
    'Dino':       'Mikanović',
    'Elio':       'Capradossi',
    'Atanas':     'Trică',
    'Bic':        'Bic',
    'Murgia':     'Murgia',
    'Chipciu':    'Chipciu',
    'Cristea':    'Cristea',
    'Codrea':     'Codrea',
    'Drammeh':    'Drammeh',
    'Fabry':      'Fábry',
    'Cisse':      'Cissé',
    'Macalou':    'Macalou',
    'Nistor':     'Nistor',
    'Simion':     'Simion',
    'Chintes':    'Chintes',
    'El Sawy':    'Sawy',
    'Postolachi': 'Postolachi',
    'Gheorghita': 'Gheorghiță',
    'Miguel':     'Miguel Silva',
    'Coubis':     'Coubis',
    # U21 / trials — no Wyscout match data, but they train with the squad.
    # We auto-create Player rows for them with is_cluj=False but in_current_squad=True.
    'Moraru':     '__create__',
    'Orban':      '__create__',
    'Bota':       '__create__',
    'Jasper':     '__create__',
    'Raji':       '__create__',
    'Taiwo':      '__create__',
}

# Stable IDs for non-Wyscout players (training only). Use 9999xx range to avoid clash with Wyscout IDs.
TRAINING_ONLY_IDS = {
    'Moraru':  9999001,
    'Orban':   9999002,
    'Bota':    9999003,
    'Jasper':  9999004,
    'Raji':    9999005,
    'Taiwo':   9999006,
}

# Friendly full names for the training-only players (used in display)
TRAINING_ONLY_NAMES = {
    'Moraru': 'M. Moraru',
    'Orban':  'A. Orban',
    'Bota':   'Bota',
    'Jasper': 'Jasper',
    'Raji':   'Raji',
    'Taiwo':  'Taiwo',
}


SESSION_TYPE_KEYWORDS = {
    'REZISTENTA': 'rezistenta',
    'FORTA':      'forta',
    'TACTIC':     'tactic',
    'REACTI':     'reactivitate',
    'AEROB':      'aerob',
    'TEHNIC':     'tehnic',
    'SSG':        'ssg',
}


DATE_RE = re.compile(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})')


def parse_session_meta(name: str):
    """Return (date_obj_or_None, session_type_or_empty)."""
    if not name:
        return None, ''
    name_upper = name.upper()
    sess_type = ''
    for kw, label in SESSION_TYPE_KEYWORDS.items():
        if kw in name_upper:
            sess_type = label
            break
    m = DATE_RE.search(name)
    date_obj = None
    if m:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            date_obj = datetime(y, mo, d).date()
        except Exception:
            pass
    return date_obj, sess_type


class Command(BaseCommand):
    help = 'Ingest GPS / wearable device training data from XLSX files.'

    def add_arguments(self, parser):
        parser.add_argument('folder', type=str, help='Folder containing *.xlsx files')
        parser.add_argument('--reset', action='store_true', help='Wipe TrainingSession before import')

    def handle(self, folder, reset=False, **opts):
        folder = Path(folder)
        if not folder.is_dir():
            self.stderr.write(f'Not a folder: {folder}')
            return

        if reset:
            self.stdout.write('Wiping existing TrainingSession rows...')
            TrainingSession.objects.all().delete()

        # Build alias -> player_obj cache (resolve once)
        cluj_players = list(Player.objects.filter(is_cluj=True))
        resolved = {}
        unresolved = set()
        for short, hint in ALIAS.items():
            if hint is None:
                resolved[short] = None
                continue
            if hint == '__create__':
                # Training-only player: auto-create or fetch
                pid = TRAINING_ONLY_IDS.get(short)
                if pid is None:
                    unresolved.add(short)
                    continue
                p, _ = Player.objects.get_or_create(
                    wy_id=pid,
                    defaults=dict(
                        name=TRAINING_ONLY_NAMES.get(short, short),
                        is_cluj=False,
                        in_current_squad=True,
                        position_code='',
                    ),
                )
                resolved[short] = p
                continue
            match = next((p for p in cluj_players if hint in p.name), None)
            resolved[short] = match
            if not match:
                unresolved.add(short)

        if unresolved:
            self.stdout.write(self.style.WARNING(
                f'Aliases not found in DB (will skip rows): {sorted(unresolved)}'
            ))

        total_imported = 0
        skipped_no_player = 0
        unknown_aliases = set()

        for xlsx in sorted(folder.glob('*.xlsx')):
            self.stdout.write(f'Reading {xlsx.name}...')
            wb = load_workbook(xlsx, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                session_name = str(row[0]).strip()
                week = row[1] if isinstance(row[1], int) else None
                short = (row[2] or '').strip()

                if short not in resolved:
                    unknown_aliases.add(short)
                    skipped_no_player += 1
                    continue
                player = resolved[short]
                if player is None:
                    skipped_no_player += 1
                    continue

                date_obj, sess_type = parse_session_meta(session_name)

                def f(idx):
                    v = row[idx]
                    try:
                        return float(v) if v is not None else 0
                    except (TypeError, ValueError):
                        return 0

                def i(idx):
                    return int(f(idx))

                TrainingSession.objects.update_or_create(
                    player=player,
                    session_name=session_name,
                    defaults=dict(
                        date=date_obj,
                        week=week,
                        session_type=sess_type,
                        duration_min=f(5),
                        distance_m=f(3),
                        work_rate=f(4),
                        accel_low_pos=i(6),
                        accel_low_neg=i(7),
                        accel_high_pos=i(8),
                        accel_high_neg=i(9),
                        high_int_acc_m=f(10),
                        high_int_dec_m=f(11),
                        speed_15_20_m=f(12),
                        speed_20_25_m=f(13),
                        speed_25_50_m=f(14),
                        power_avg_wkg=f(15),
                        sprints_per_min=f(16),
                    ),
                )
                total_imported += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done. Imported/updated {total_imported} sessions. Skipped {skipped_no_player} rows (unmapped or unknown players).'
        ))
        if unknown_aliases:
            self.stdout.write(f'Aliases NOT in mapping at all: {sorted(unknown_aliases)}')
        # Per-player summary
        from django.db.models import Count
        per_player = TrainingSession.objects.values('player__name').annotate(n=Count('id')).order_by('-n')
        self.stdout.write('Top 5 by session count:')
        for r in per_player[:5]:
            self.stdout.write(f"  {r['player__name']}: {r['n']}")
